from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from .schemas import TeachingClassGroup
from .workflow import TEACHING_CLASS_COLUMNS, flatten_groups


def teaching_class_to_xlsx(
    groups: list[TeachingClassGroup],
    *,
    sheet_name: str = "教学班清洗结果",
) -> bytes:
    output = flatten_groups(groups).fillna("")
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        output.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [16, 22, 18, 16, 18, 18, 20, 20]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = width
        for row in worksheet.iter_rows(min_row=2, max_col=len(TEACHING_CLASS_COLUMNS)):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
    return buffer.getvalue()
